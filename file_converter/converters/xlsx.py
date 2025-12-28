"""
file_converter/converters/xlsx.py
XLSX 文件转换器 - 使用 openpyxl 提取表格内容
"""

import io
import logging
from typing import BinaryIO, List, Optional

from openpyxl import load_workbook

from file_converter.converters.base import BaseConverter, ConversionResult

logger = logging.getLogger(__name__)


class XlsxConverter(BaseConverter):
    """
    XLSX 转 Markdown 转换器
    
    功能：
    - 每个 Sheet 转换为独立区块，以 "## SheetName" 分隔
    - 第一行作为表头
    - 生成 Markdown 表格语法
    - 支持合并单元格（内容放在左上角）
    """
    
    supported_extensions = ["xlsx", "xls"]
    
    # 最大行列限制，避免处理超大文件
    MAX_ROWS = 1000
    MAX_COLS = 50
    
    def convert(self, file: BinaryIO, filename: str) -> ConversionResult:
        """将 XLSX 转换为 Markdown"""
        warnings: List[str] = []
        content_parts: List[str] = []
        
        try:
            wb = load_workbook(file, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            logger.info(f"Processing XLSX: {filename}, {len(sheet_names)} sheets")
            
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                sheet_content = self._process_sheet(sheet, sheet_name, warnings)
                if sheet_content:
                    content_parts.append(sheet_content)
            
            wb.close()
        
        except Exception as e:
            logger.error(f"XLSX conversion failed: {filename}, error: {e}")
            raise ValueError(f"XLSX 转换失败: {str(e)}")
        
        content = "\n\n".join(content_parts)
        
        if not content.strip():
            warnings.append("XLSX 未提取到任何内容")
        
        return ConversionResult(content=content, warnings=warnings)
    
    def _process_sheet(self, sheet, sheet_name: str, warnings: List[str]) -> str:
        """处理单个工作表"""
        parts: List[str] = []
        
        # 添加 Sheet 标题
        parts.append(f"## {sheet_name}")
        
        # 获取有效数据范围
        rows = list(sheet.iter_rows(max_row=self.MAX_ROWS, max_col=self.MAX_COLS))
        
        if not rows:
            return ""
        
        # 查找实际数据范围（跳过空行）
        data_rows: List[List[str]] = []
        max_col = 0
        
        for row in rows:
            cells = [self._get_cell_value(cell) for cell in row]
            
            # 检查是否是空行
            if not any(cells):
                continue
            
            # 更新最大列数
            for i, cell in enumerate(cells):
                if cell:
                    max_col = max(max_col, i + 1)
            
            data_rows.append(cells)
        
        if not data_rows or max_col == 0:
            return ""
        
        # 检查是否超过限制
        if len(data_rows) > self.MAX_ROWS:
            warnings.append(f"Sheet '{sheet_name}' 行数超过 {self.MAX_ROWS}，已截断")
            data_rows = data_rows[:self.MAX_ROWS]
        
        # 转换为 Markdown 表格
        table_md = self._rows_to_markdown(data_rows, max_col)
        if table_md:
            parts.append(table_md)
        
        return "\n\n".join(parts)
    
    def _get_cell_value(self, cell) -> str:
        """获取单元格值"""
        if cell.value is None:
            return ""
        
        value = str(cell.value).strip()
        
        # 转义管道符和换行
        value = value.replace("|", "\\|")
        value = value.replace("\n", " ")
        
        return value
    
    def _rows_to_markdown(self, rows: List[List[str]], col_count: int) -> str:
        """将行数据转换为 Markdown 表格"""
        if not rows:
            return ""
        
        lines: List[str] = []
        
        for row_idx, row in enumerate(rows):
            # 确保每行有足够的列
            cells = row[:col_count]
            while len(cells) < col_count:
                cells.append("")
            
            lines.append("| " + " | ".join(cells) + " |")
            
            # 在第一行后添加分隔行
            if row_idx == 0:
                lines.append("| " + " | ".join(["---"] * col_count) + " |")
        
        return "\n".join(lines)
